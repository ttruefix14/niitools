import arcpy
import os
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font

def execute():
    """The source code of the tool."""
    arcpy.env.addOutputsToMap = True
    
    # Задаем параметры
    params = arcpy.GetParameterInfo()
    deleteMemory = params[0]
    adjustValues = params[1]
    mo_type = params[2]
    inputXLS = params[3]
    layer_AdmeMO = params[4]
    sql_AdmeMO = params[5]
    layer_AdmeNP = params[6]
    field_AdmeNP = params[7]
    area_field_AdmeNP = params[8]
    layer_LandUse = params[9]
    layer_SI = params[10]
    field_SI = params[11]
    layer_FZ = params[12]
    field_FZ = params[13]
    outputXLS = params[14]

    mo_dict = {"Сельское поселение": "Функциональные зоны сельского поселения", 
               "Городское поселение": "Функциональные зоны городского поселения", 
               "Городской округ": "Функциональные зоны городского округа"}
    mo_type = mo_dict[mo_type.valueAsText]
    
    # Задаем папку проекта, в которой будут создаваться файлы
    defaultFolder = arcpy.mp.ArcGISProject("CURRENT").homeFolder
    
    # Удалям 'memory' или создаем базу данных с временными файлами
    if deleteMemory.value:
        output_db = r"memory" #os.path.join(defaultFolder, "temp.gdb")
        arcpy.management.Delete(output_db)
    else:
        arcpy.management.CreateFileGDB(defaultFolder, "temp.gdb")
        output_db = os.path.join(defaultFolder, "temp.gdb")
    
    ### Выгружаем слои в локальную базу, чтобы переименовать столбцы
    arcpy.conversion.FeatureClassToFeatureClass(layer_AdmeMO.valueAsText, output_db, "AdmeMO", sql_AdmeMO.value)

    arcpy.conversion.FeatureClassToFeatureClass(layer_AdmeNP.valueAsText, output_db, "AdmeNP")

    if field_AdmeNP.valueAsText != "Note":
        if "Note" in [field.name for field in arcpy.ListFields(os.path.join(output_db, "AdmeNP"))]:
            arcpy.management.DeleteField(os.path.join(output_db, "AdmeNP"), ['Note'])

        arcpy.management.AddField(os.path.join(output_db, "AdmeNP"), "Note", "TEXT")
        arcpy.management.CalculateField(os.path.join(output_db, "AdmeNP"), "Note", f"!{field_AdmeNP.valueAsText}!")
        # arcpy.management.AlterField(os.path.join(output_db, "AdmeNP"), field_AdmeNP.valueAsText, "Note", "Note")

    arcpy.conversion.FeatureClassToFeatureClass(layer_LandUse.valueAsText, output_db, "LandUse")
    arcpy.conversion.FeatureClassToFeatureClass(layer_SI.valueAsText, output_db, "FunctionalZone_si")
    arcpy.conversion.FeatureClassToFeatureClass(layer_FZ.valueAsText, output_db, "FunctionalZone")

    if field_SI.valueAsText != "Ext_Zone_Code":
        if "Ext_Zone_Code" in [field.name for field in arcpy.ListFields(os.path.join(output_db, "FunctionalZone_si"))]:
            arcpy.management.DeleteField(os.path.join(output_db, "FunctionalZone_si"), [field_SI.valueAsText])

        arcpy.management.AddField(os.path.join(output_db, "FunctionalZone_si"), "Ext_Zone_Code", "TEXT")
        arcpy.management.CalculateField(os.path.join(output_db, "FunctionalZone_si"), "Ext_Zone_Code", f"!{field_SI.valueAsText}!")
        # arcpy.management.AlterField(os.path.join(output_db, "FunctionalZone_si"), field_SI.valueAsText, "Ext_Zone_Code", "Ext_Zone_Code")

    if field_FZ.valueAsText != "Ext_Zone_Code":
        if "Ext_Zone_Code" in [field.name for field in arcpy.ListFields(os.path.join(output_db, "FunctionalZone"))]:
            arcpy.management.DeleteField(os.path.join(output_db, "FunctionalZone"), [field_FZ.valueAsText])

        arcpy.management.AddField(os.path.join(output_db, "FunctionalZone"), "Ext_Zone_Code", "TEXT")
        arcpy.management.CalculateField(os.path.join(output_db, "FunctionalZone"), "Ext_Zone_Code", f"!{field_FZ.valueAsText}!")
        # arcpy.management.AlterField(os.path.join(output_db, "FunctionalZone"), field_FZ.valueAsText, "Ext_Zone_Code", "Ext_Zone_Code")
    
    # Подгружаем эксель с входными данными
    xls = pd.ExcelFile(inputXLS.valueAsText)

    NP_type = pd.read_excel(xls, "Типы НП")
    NP = pd.read_excel(xls, "Населенные пункты")
    zones = pd.read_excel(xls, "Функциональные зоны")
    xl = pd.read_excel(xls, "Категории земель")

    ###ARCGIS TABLE TO PANDAS DF###
    def table_to_data_frame(in_table, input_fields=None, where_clause=None):
        """Function will convert an arcgis table into a pandas dataframe with an object ID index, and the selected
        input fields using an arcpy.da.SearchCursor."""
        OIDFieldName = arcpy.Describe(in_table).OIDFieldName
        try:
            shapeFieldName = arcpy.Describe(in_table).shapeFieldName
        except:
            shapeFieldName = None
        if input_fields:
            if shapeFieldName:
                final_fields = [OIDFieldName] + ['SHAPE@'] + input_fields
            else:
                final_fields = [OIDFieldName] + input_fields
        else:
            final_fields = [field.name if field.name != shapeFieldName else 'SHAPE@' for field in arcpy.ListFields(in_table)]
        data = [row for row in arcpy.da.SearchCursor(in_table, final_fields, where_clause=where_clause)]
        fc_dataframe = pd.DataFrame(data, columns=final_fields)
        fc_dataframe = fc_dataframe.set_index(OIDFieldName, drop=True)
        return fc_dataframe

    ### Добавляем слои существующих и планирумых населенных пунктов
    aprx = arcpy.mp.ArcGISProject("CURRENT")
    m = aprx.listMaps(aprx.activeMap.name)[0]   ### НАЗВАНИЕ КАРТЫ

    for lyr in m.listLayers():
        try:
            if lyr.name in ["AdmeNP_SI", "AdmeNP_PLAN", "AdmeNP_all", "LandUse_si", "LandUse_plan"]:
                arcpy.management.Delete(lyr)
        except:
            pass
    
    arcpy.management.MakeFeatureLayer(output_db+"\AdmeNP", "AdmeNP_SI", "STATUS_ADM = 1")

    arcpy.management.MakeFeatureLayer(output_db+"\AdmeNP", "AdmeNP_PLAN", "STATUS_ADM = 2")

    arcpy.management.MakeFeatureLayer(output_db+"\AdmeNP", "AdmeNP_all", "Note IN ('не изменяемая', 'планируемая')")
    
    ###Добавляем слои существующих и планируемых категорий
    arcpy.env.workspace = "CURRENT"
    arcpy.management.MakeFeatureLayer(output_db+"\LandUse", "LandUse_si", "STATUS = 1")

    arcpy.management.MakeFeatureLayer(output_db+"\LandUse", "LandUse_plan", "STATUS = 2")

    ###Находим пересечения между существующими и планируемыми категориями
    arcpy.analysis.PairwiseIntersect(('LandUse_si', 'LandUse_plan'), output_db + '\LU_int', "all", "", "input")
    ###Находим площади план ФЗ
    arcpy.analysis.PairwiseIntersect((output_db + "\FunctionalZone", 'AdmeNP_all'), output_db + "\FZ_NP1", "all", "", "input")
    arcpy.analysis.Erase(output_db + "\FunctionalZone", 'AdmeNP_all', output_db + '\FZ_MO1', "")
    ###Находим площади сущ ФЗ
    arcpy.analysis.PairwiseIntersect((output_db + "\FunctionalZone_si", "AdmeNP_SI"), output_db + "\SI_NP1", "all", "", "input")
    arcpy.analysis.Erase(output_db + "\FunctionalZone_si", 'AdmeNP_SI', output_db + '\SI_MO1', "")
    ### Находим площади иных категорий в границах сущ НП
    arcpy.analysis.PairwiseIntersect((output_db + "\LandUse", "AdmeNP_SI"), output_db + "\LU_MO", "all", "", "input")
    ### Находим земли НП за границами сущ НП
    arcpy.analysis.Erase(output_db + "\LandUse", 'AdmeNP_SI', output_db + '\LU_NP', "")
    ### Вносим поправку на пересекающиеся дублирующиеся категории с планируемыми категориями
    try:
        arcpy.analysis.Intersect((output_db + '\LU_int'), output_db + '\LU_int_minus', "all", "", "input")
    except:
        arcpy.conversion.FeatureClassToFeatureClass(output_db + '\LU_int', output_db, 'LU_int_minus')
        arcpy.management.DeleteFeatures(output_db + '\LU_int_minus')
    ### Находим площади планируемых категорий на землях НП за границами НП
    arcpy.analysis.PairwiseIntersect((output_db + '\LU_NP', "LandUse_plan"), output_db + "\LU_NP_minus", "all", "", "input")
    
    ### Диссолв слоев зонирования
    arcpy.management.Dissolve(output_db + "\SI_NP1", output_db + '\SI_NP', ['NAME', 'SETTL_TYPE', 'Ext_Zone_Code'])
    arcpy.management.Dissolve(output_db + "\FZ_NP1", output_db + '\FZ_NP', ['NAME', 'SETTL_TYPE', 'Ext_Zone_Code'])
    arcpy.management.Dissolve(output_db + "\SI_MO1", output_db + '\SI_MO', ['Ext_Zone_Code'])
    arcpy.management.Dissolve(output_db + "\FZ_MO1", output_db + '\FZ_MO', ['Ext_Zone_Code'])

    
    ### Находим площади земель нп за границами НП, включенных в границы НП
    arcpy.analysis.Erase('AdmeNP_PLAN', 'AdmeNP_SI', output_db + '\\NP_vkl', "")
    arcpy.analysis.PairwiseIntersect(('LandUse_si', output_db + '\\NP_vkl'), output_db + '\LU_NP_vkl', "all", "", "input")
    
    ### Находим площади иных категорий в границах НП, исключенных из границ НП
    arcpy.analysis.Erase('AdmeNP_SI', 'AdmeNP_all', output_db + '\\NP_iskl', "")
    arcpy.analysis.PairwiseIntersect(('LandUse_si', output_db + '\\NP_iskl'), output_db + '\LU_MO_plan', "all", "", "input")


    FZ_NP = table_to_data_frame(os.path.join(output_db, 'FZ_NP'), ['NAME', 'SETTL_TYPE', 'Ext_Zone_Code'])
    # FZ_NP['Shape_Area'] = FZ_NP.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    FZ_NP['Shape_Area'] = FZ_NP.apply(lambda row: row["SHAPE@"].area if row["SHAPE@"].area > 500 else (501 if row["SHAPE@"].area > 100 else 0), axis=1, result_type='reduce')
    SI_NP = table_to_data_frame(os.path.join(output_db, 'SI_NP'), ['NAME', 'SETTL_TYPE', 'Ext_Zone_Code'])
    SI_NP['Shape_Area'] = SI_NP.apply(lambda row: row["SHAPE@"].area if row["SHAPE@"].area > 500 else (501 if row["SHAPE@"].area > 100 else 0), axis=1, result_type='reduce')
    FZ_MO = table_to_data_frame(os.path.join(output_db, 'FZ_MO'), ['Ext_Zone_Code'])
    FZ_MO['Shape_Area'] = FZ_MO.apply(lambda row: row["SHAPE@"].area if row["SHAPE@"].area > 500 else (501 if row["SHAPE@"].area > 100 else 0), axis=1, result_type='reduce')
    SI_MO = table_to_data_frame(os.path.join(output_db, 'SI_MO'), ['Ext_Zone_Code'])
    SI_MO['Shape_Area'] = SI_MO.apply(lambda row: row["SHAPE@"].area if row["SHAPE@"].area > 500 else (501 if row["SHAPE@"].area > 100 else 0), axis=1, result_type='reduce')

    # Задаем определенное поле для площади НП
    fields_AdmeNP = ['NAME', 'SETTL_TYPE', 'NOTE']
    if area_field_AdmeNP.value:
        fields_AdmeNP.append(area_field_AdmeNP.valueAsText)
    AdmeNP = table_to_data_frame(os.path.join(output_db, 'AdmeNP'), fields_AdmeNP)
    if area_field_AdmeNP.value:
        AdmeNP['Shape_Area'] = AdmeNP[area_field_AdmeNP.valueAsText]
    else:
        AdmeNP['Shape_Area'] = AdmeNP.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')

    AdmeMO = table_to_data_frame(os.path.join(output_db, 'AdmeMO'), ['STATUS_ADM'])
    AdmeMO['Shape_Area'] = AdmeMO.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    LU_S = table_to_data_frame(os.path.join(output_db, 'LandUse'), ['CLASSID', 'STATUS'])
    LU_S['Shape_Area'] = LU_S.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    LU_int = table_to_data_frame(os.path.join(output_db, 'LU_int'), ['CLASSID', 'CLASSID_1'])
    LU_int['Shape_Area'] = LU_int.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    LU_NP = table_to_data_frame(os.path.join(output_db, 'LU_NP'), ['CLASSID', 'STATUS'])
    LU_NP['Shape_Area'] = LU_NP.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    LU_MO = table_to_data_frame(os.path.join(output_db, 'LU_MO'), ['CLASSID', 'STATUS', 'NAME'])
    LU_MO['Shape_Area'] = LU_MO.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    LU_MO_plan = table_to_data_frame(os.path.join(output_db, 'LU_MO_plan'), ['CLASSID', 'STATUS', 'NAME'])
    LU_MO_plan['Shape_Area'] = LU_MO_plan.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    LU_NP_vkl = table_to_data_frame(os.path.join(output_db, 'LU_NP_vkl'), ['CLASSID'])
    LU_NP_vkl['Shape_Area'] = LU_NP_vkl.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    LU_NP_minus = table_to_data_frame(os.path.join(output_db, "LU_NP_minus"), ['CLASSID', 'CLASSID_1', 'STATUS'])
    LU_NP_minus['Shape_Area'] = LU_NP_minus.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    LU_int_minus = table_to_data_frame(os.path.join(output_db, 'LU_int_minus'), ['CLASSID', 'CLASSID_1'])
    LU_int_minus['Shape_Area'] = LU_int_minus.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    NP_vkl = table_to_data_frame(os.path.join(output_db + '\\NP_vkl'))
    NP_vkl['Shape_Area'] = NP_vkl.apply(lambda row: row["SHAPE@"].area, axis=1, result_type='reduce')
    NP_vkl['Area'] = NP_vkl['Shape_Area']/10000
    FZ_NP['SI+PLAN'], SI_NP['SI'], FZ_MO['SI+PLAN'], SI_MO['SI'], LU_int['Area'], LU_S['Area']  = FZ_NP['Shape_Area']/10000, SI_NP['Shape_Area']/10000, FZ_MO['Shape_Area']/10000, SI_MO['Shape_Area']/10000, LU_int['Shape_Area']/10000, LU_S['Shape_Area']/10000
    AdmeMO['Area'] = AdmeMO['Shape_Area']/10000
    LU_NP['Area'] = LU_NP['Shape_Area']/10000
    LU_MO['Area'] = LU_MO['Shape_Area']/10000
    LU_MO_plan['Area'] = LU_MO_plan['Shape_Area']/10000
    LU_NP_vkl['Area'] = LU_NP_vkl['Shape_Area']/10000
    LU_NP_minus['Area'] = LU_NP_minus['Shape_Area']/10000
    LU_int_minus['Area'] = LU_int_minus['Shape_Area']/10000

    FZ_NP = FZ_NP.round(1)
    SI_NP = SI_NP.round(1)
    FZ_MO = FZ_MO.round(1)
    SI_MO = SI_MO.round(1)
    AdmeNP = AdmeNP.round(1)



    # NP_type = pd.read_excel(io = os.path.join(defaultFolder, "NP_TYPE.xlsx"))
    SI_NP = SI_NP.merge(NP_type, left_on='SETTL_TYPE', right_on = 'SETTL_TYPE', how = 'left')
    SI_NP['NAME'] = SI_NP['NP_TYPE'] + ' ' + SI_NP['NAME']
    FZ_NP = FZ_NP.merge(NP_type, left_on='SETTL_TYPE', right_on = 'SETTL_TYPE', how = 'left')
    FZ_NP['NAME'] = FZ_NP['NP_TYPE'] + ' ' + FZ_NP['NAME']
    ##
    FZ_MO['NAME'] = 'МО'
    SI_MO['NAME'] = 'МО'
    FZ_NP = FZ_NP.append(FZ_MO, ignore_index=False, verify_integrity=False, sort=False)
    SI_NP = SI_NP.append(SI_MO, ignore_index=False, verify_integrity=False, sort=False)
    ##
    AdmeNP['Area'] = AdmeNP['Shape_Area']/10000
    AdmeNP = AdmeNP.merge(NP_type, left_on='SETTL_TYPE', right_on = 'SETTL_TYPE', how = 'left')
    AdmeNP['NAME'] = AdmeNP['NP_TYPE'] + ' ' + AdmeNP['NAME']

    AdmeNP_S = AdmeNP.loc[AdmeNP['NOTE'] != 'планируемая']
    AdmeNP_P = AdmeNP.loc[AdmeNP['NOTE'] != 'изменяемая']
    noNP = list(set(AdmeNP_P['NAME'].to_list()) - set(AdmeNP_S['NAME'].to_list()))
    if len(noNP) > 0:
        df_noNP = pd.DataFrame([[i, 0] for i in noNP], columns=['NAME', 'AREA'])
        AdmeNP_S = AdmeNP_S.append(df_noNP, ignore_index=False, verify_integrity=False, sort=False)

    # zones = pd.read_excel(io = os.path.join(defaultFolder, "zones.xlsx"))
    SI_NP = SI_NP.merge(zones, left_on='Ext_Zone_Code', right_on = 'Ext_Zone_Code', how = 'left')
    FZ_NP = FZ_NP.merge(zones, left_on='Ext_Zone_Code', right_on = 'Ext_Zone_Code', how = 'left')
    SI_NP = SI_NP[['Ext_Zone_Code', 'NAME', 'Zone_group', 'SI']]
    FZ_NP = FZ_NP[['Ext_Zone_Code', 'NAME', 'Zone_group', 'SI+PLAN']]
    zone_group = []
    # podzone_group = []
    q = 0

    for group in range (len(zones['Zone_group'])):
        if q != 0:
            if zones['Zone_group'][group] != zones['Zone_group'][group-1]:
                zone_group += [zones['Zone_group'][group]]
        elif zones['Zone_group'][group] != 1:
            zone_group += [zones['Zone_group'][group]]
            q += 1
    print(zone_group)
    ##    q = 0
    ##    for group in range (len(zones['Podzone_group'])-1):
    ##        if q != 0:
    ##            if zones['Podzone_group'][group] != zones['Podzone_group'][group-1]:
    ##                podzone_group += [zones['Podzone_group'][group]]
    ##        elif zones['Podzone_group'][group] != 1:
    ##            podzone_group += [zones['Podzone_group'][group]]
    ##            q += 1
    ##    print(podzone_group)
    zone_index = []
    for zone in zones['Ext_Zone_Code']:
        if zone != 0:
            zone_index += [zone]
    print(zone_index)

    # NP = pd.read_excel(io = os.path.join(defaultFolder, "NP.xlsx"))
    q = 0
    okruga = []
    for okrug in range (len(NP['OKRUG'])):
        if q != 0:
            if NP['OKRUG'][okrug] != NP['OKRUG'][okrug-1]:
                okruga += [NP['OKRUG'][okrug]]
        else:
            okruga += [NP['OKRUG'][okrug]]
            q += 1
    print(okruga)
    AdmeNP_S = AdmeNP_S.merge(NP, left_on='NAME', right_on = 'NAME', how = 'left')
    AdmeNP_P = AdmeNP_P.merge(NP, left_on='NAME', right_on = 'NAME', how = 'left')

    ##
    MO = pd.DataFrame(['МО'], columns=['NAME'])
    NP = NP.append(MO, ignore_index=False, verify_integrity=False, sort=False)
    ##
    FZ_NP['SI+PLAN'] = FZ_NP['SI+PLAN'].fillna(0)
    SI_NP['SI'] =  SI_NP['SI'].fillna(0)
    ### ЛОК ПО НУЛЮ ДЛЯ СОЗДАНИЯ НОВОЙ ТАБЛИЦЫ С НУЖНЫМИ СТОЛБЦАМИ ###

    SI_FZ = pd.DataFrame(columns=['Ext_Zone_Code'])
    Balance = pd.DataFrame(columns=['Ext_Zone_Code'])
    Categories = pd.DataFrame(columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    ###   ###
    areaMO_S = float(AdmeMO.loc[AdmeMO['STATUS_ADM'] == 1, ['Area']].sum())
    if float(AdmeMO.loc[AdmeMO['STATUS_ADM'] == 2, ['Area']].sum()) != 0:
        areaMO_P = float(AdmeMO.loc[AdmeMO['STATUS_ADM'] == 2, ['Area']].sum())
    else:
        areaMO_P = float(AdmeMO.loc[AdmeMO['STATUS_ADM'] == 1, ['Area']].sum())

    block = pd.DataFrame([['1.', 'Территория поселения всего', areaMO_S, areaMO_P]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    Balance = Balance.append(block, ignore_index=False, verify_integrity=False, sort=False)
    ###
    ###  СУММА КАТЕГОРИЙ ###
    ###
    ### Вставляем пустую строку
    #block = pd.DataFrame([[None, None, None, None]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    #Balance = Balance.append(block, ignore_index=False, verify_integrity=False, sort=False)

    LU_S = LU_S.loc[LU_S['STATUS'] == 1]
    LU_S = LU_S.loc[LU_S['CLASSID'] != 702010100]
    ### КАТЕГОРИИ ###
    # xl = pd.read_excel(io = os.path.join(defaultFolder, "Категории.xlsx"))
    LU_S = LU_S.merge(xl, left_on='CLASSID', right_on = 'CLASSID', how = 'left')
    LU_int = LU_int.merge(xl, left_on='CLASSID', right_on = 'CLASSID', how = 'left')
    LU_int = LU_int.merge(xl, left_on='CLASSID_1', right_on = 'CLASSID', how = 'left')

    LU_int_minus = LU_int_minus.rename(columns={'CLASSID': 'CLASSID_old'})
    LU_int_minus = LU_int_minus.merge(xl, left_on='CLASSID_1', right_on = 'CLASSID', how = 'left')
    ##земли нп за границами нп
    LU_NP = LU_NP.loc[LU_NP['STATUS'] == 1]
    LU_NP = LU_NP.loc[LU_NP['CLASSID'] == 702010100]
    LU_NP = LU_NP.merge(xl, left_on='CLASSID', right_on = 'CLASSID', how = 'left')

    LU_NP_minus = LU_NP_minus.loc[LU_NP_minus['STATUS'] == 1]
    LU_NP_minus = LU_NP_minus.loc[LU_NP_minus['CLASSID'] == 702010100]
    LU_NP_minus = LU_NP_minus.merge(xl, left_on='CLASSID_1', right_on = 'CLASSID', how = 'left')

    LU_NP_vkl = LU_NP_vkl.loc[LU_NP_vkl['CLASSID'] == 702010100]
    ###земли иных категорий в границах нп
    LU_MO = LU_MO.loc[LU_MO['STATUS'] == 1]
    LU_MO = LU_MO.loc[LU_MO['CLASSID'] != 702010100]
    LU_MO = LU_MO.merge(xl, left_on='CLASSID', right_on = 'CLASSID', how = 'left')
    LU_MO_plan = LU_MO_plan.loc[LU_MO_plan['STATUS'] == 1]
    LU_MO_plan = LU_MO_plan.loc[LU_MO_plan['CLASSID'] != 702010100]
    LU_MO_plan = LU_MO_plan.merge(xl, left_on='CLASSID', right_on = 'CLASSID', how = 'left')
    ###
    #сводная для движения земли
    LU_PIV = pd.pivot_table(LU_int, index = 'CATEGORY_x', columns = 'CATEGORY_y', values = 'Area', aggfunc = np.sum)
    #
    # cats = ['Земли сельскохозяйственного назначения', 
    #         'Земли промышленности, энергетики, транспорта, связи, радиовещания, телевидения, информатики, земли для обеспечения космической деятельности, земли обороны, безопасности и земли иного специального назначения', 

    #         'Земли особо охраняемых территорий и объектов', 
    #         'Земли лесного фонда', 
    #         'Земли лесного фонда (согласно ЕГРН)',
    #         'Земли водного фонда', 
    #         'Земли запаса']
    cats = xl['CATEGORY'].to_list()[1:]

    lu_columns = []
    for cat in xl['CATEGORY'].to_list():
        if cat == 'Земли населенных пунктов':
            if float(NP_vkl['Area'].sum()) > 0.05:
                lu_columns.append(cat)
        elif cat in LU_int['CATEGORY_y'].to_list():
            lu_columns.append(cat)

    if 'Земли населенных пунктов' in LU_int['CATEGORY_x'].tolist():
        # catAreaS = float(LU_S.loc[LU_S['CATEGORY'] == cat, ['Area']].sum())
        LU_NP_to = LU_int.loc[LU_int['CATEGORY_x'] == 'Земли населенных пунктов']
        current_dict = dict.fromkeys(lu_columns, 0)
        for cat in list(dict.fromkeys(LU_NP_to['CATEGORY_y']).keys()):
            current = LU_NP_to.loc[LU_NP_to['CATEGORY_y'] == cat]
            if len(current) > 0:
                current_dict[cat] = current['Area'].sum()
    else:
        current_dict = dict.fromkeys(lu_columns, 0)

# КОСТЫЫЫЫЫЫЫЛЬ
    LU_NP_to = LU_NP_minus
    current_dict_minus = dict.fromkeys(lu_columns, 0)
    for cat in list(dict.fromkeys(LU_NP_to['CATEGORY']).keys()):
        current = LU_NP_to.loc[LU_NP_to['CATEGORY'] == cat]
        if len(current) > 0:
            
            current_dict_minus[cat] = current['Area'].sum()
    
    current_dict = {key: value - current_dict_minus.get(key, 0) for key, value in current_dict.items()}

    arcpy.AddMessage("lu columns: " + str(lu_columns))
    arcpy.AddMessage("current_dict: " + str(current_dict))

    if "Земли населенных пунктов" in lu_columns and "Земли населенных пунктов" not in current_dict:
        raise Exception("Границы НП не расширяются, однако присутствует планируемая категория НП")

    catNP = pd.DataFrame([['1.1', 'Земли населенных пунктов', AdmeNP_S['Area'].sum(), AdmeNP_P['Area'].sum()]+list(current_dict.values())+[None]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN']+lu_columns+['ИТОГО'])
    Balance = Balance.append(catNP, ignore_index=False, verify_integrity=False, sort=False)

    for okr in okruga:
        AdmeNP_SA = AdmeNP_S.loc[AdmeNP_S['OKRUG'] == okr]
        AdmeNP_PA = AdmeNP_P.loc[AdmeNP_P['OKRUG'] == okr]
        #okr_name = pd.DataFrame([[None, okr, AdmeNP_SA['Area'].sum(), AdmeNP_PA['Area'].sum()]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
        okr_name = pd.DataFrame([[None, okr, None, None]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
        Balance = Balance.append(okr_name, ignore_index=False, verify_integrity=False, sort=False)
        for name in NP['NAME']:
            if name not in AdmeNP_SA['NAME'].tolist():   ####ПОПРАВИТЬ НА СЛУЧАЙ ЕСЛИ БУДЕТ ПОЯВЛЯТЬСЯ НОВЫЙ НАСЕЛЕННЫЙ ПУНКТ
                continue
            AdmeNP_SB = AdmeNP_SA.loc[AdmeNP_SA['NAME'] == name]
            AdmeNP_PB = AdmeNP_PA.loc[AdmeNP_PA['NAME'] == name]
            np_name = pd.DataFrame([[None, name, AdmeNP_SB['Area'].sum(), AdmeNP_PB['Area'].sum()]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
            Balance = Balance.append(np_name, ignore_index=False, verify_integrity=False, sort=False)

    o = 2
    catSUM_SA = 0
    catSUM_PA = 0
    for cat in cats:
        if cat in LU_S['CATEGORY'].tolist() or cat in LU_int['CATEGORY_y'].tolist():
            if cat in LU_int['CATEGORY_x'].tolist():
                # catAreaS = float(LU_S.loc[LU_S['CATEGORY'] == cat, ['Area']].sum())
                LU_NP_to = LU_int.loc[LU_int['CATEGORY_x'] == cat]
                current_dict = dict.fromkeys(lu_columns, 0)
                for cat2 in list(dict.fromkeys(LU_NP_to['CATEGORY_y']).keys()):
                    #CHEEECK
                    if cat2 not in lu_columns:
                        continue
                    current = LU_NP_to.loc[LU_NP_to['CATEGORY_y'] == cat2]
                    if len(current) > 0:
                        current_dict[cat2] = current['Area'].sum()
            else:
                current_dict = dict.fromkeys(lu_columns, 0) #current_dict(dict.fromkeys(current_dict.keys(), 0))
                        
            catAreaS = float(LU_S.loc[LU_S['CATEGORY'] == cat, ['Area']].sum())
            catAreaP = float(LU_int.loc[LU_int['CATEGORY_y'] == cat, ['Area']].sum()) - float(LU_int.loc[LU_int['CATEGORY_x'] == cat, ['Area']].sum()) - float(LU_int_minus.loc[LU_int_minus['CATEGORY'] == cat, ['Area']].sum())/2
            
            arcpy.AddMessage("lu_columns: " + str(lu_columns))

            arcpy.AddMessage("current_dict: " + str(current_dict))
            
            block = pd.DataFrame([['1.'+ str(o), cat, catAreaS, catAreaP]+list(current_dict.values())+[None]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'PLAN']+lu_columns+['ИТОГО'])
            Categories = Categories.append(block, ignore_index=False, verify_integrity=False, sort=False)
            o += 1
            catSUM_SA += catAreaS
            catSUM_PA += (catAreaP + catAreaS)

    #### ДОПИСАТЬ КОД ПО ПЛАНИРУЕМЫМ КУСКАМ НП ИНТЕРСЕКТ КАТЕГОРИЙ ДЛЯ ВЫВОДА ПЛОЩАДИ ЭТИХ НП ЗА НП
    catAreaNP = float(LU_NP['Area'].sum())
    ### ВЫЧЕСТЬ СУММУ ПЛАН КАТЕГОРИЙ


    # catAreaS = float(LU_S.loc[LU_S['CATEGORY'] == cat, ['Area']].sum())
    LU_NP_to = LU_NP_minus
    current_dict = dict.fromkeys(lu_columns, 0)
    for cat in list(dict.fromkeys(LU_NP_to['CATEGORY']).keys()):
        current = LU_NP_to.loc[LU_NP_to['CATEGORY'] == cat]
        if len(current) > 0:
            current_dict[cat] = current['Area'].sum()

    if float(LU_NP_vkl['Area'].sum()) > 0.05:
        current_dict['Земли населенных пунктов'] = float(LU_NP_vkl['Area'].sum())
    catAreaNP_plan = float(LU_NP_vkl['Area'].sum()) + float(LU_NP_minus['Area'].sum())
    if catAreaNP > 0.1 or catAreaNP_plan > 0.1:
        arcpy.AddMessage(str(current_dict))
        block = pd.DataFrame([['1.'+ str(o), 'Земли населенных пунктов за границами населенных пунктов, сведения о которых внесены в ЕГРН', catAreaNP, -catAreaNP_plan]+list(current_dict.values()) + [None]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'PLAN']+lu_columns+['ИТОГО'])
        Categories = Categories.append(block, ignore_index=False, verify_integrity=False, sort=False)
        
    Categories['SI+PLAN'] = Categories['SI'] + Categories['PLAN']
    catSUM_S = AdmeNP_S['Area'].sum() + catAreaNP + catSUM_SA
    catSUM_P = AdmeNP_P['Area'].sum() + catAreaNP + catSUM_PA - catAreaNP_plan
    block = pd.DataFrame([['ИТОГО', None, catSUM_S, catSUM_P, catSUM_P - catSUM_S]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN', 'PLAN'])
    Categories = Categories.append(block, ignore_index=False, verify_integrity=False, sort=False)
    block = pd.DataFrame([['ошибка', None, 111111, 111111, 111111]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN', 'PLAN'])
    Categories = Categories.append(block, ignore_index=False, verify_integrity=False, sort=False)
    ###иные категории в границах считаем
    otherSUM_S = 0
    otherSUM_PLAN = 0
    catAreaS = 0
    catAreaP = 0
    for cat in cats:
        if cat in LU_MO['CATEGORY'].tolist() or cat in LU_MO_plan['CATEGORY'].tolist():
            catAreaS = float(LU_MO.loc[LU_MO['CATEGORY'] == cat, ['Area']].sum())
            catAreaP = float(LU_MO_plan.loc[LU_MO_plan['CATEGORY'] == cat, ['Area']].sum())
            otherSUM_S += catAreaS
            otherSUM_PLAN += catAreaP
    block = pd.DataFrame([['*', 'Иные категории земель в границах населенных пунктов', otherSUM_S, otherSUM_S - otherSUM_PLAN, -otherSUM_PLAN]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN', 'PLAN'])
    Categories = Categories.append(block, ignore_index=False, verify_integrity=False, sort=False)
    catAreaS = 0
    catAreaP = 0
    for cat in cats:
        if cat in LU_MO['CATEGORY'].tolist() or cat in LU_MO_plan['CATEGORY'].tolist():
            catAreaS = float(LU_MO.loc[LU_MO['CATEGORY'] == cat, ['Area']].sum())
            catAreaP = float(LU_MO_plan.loc[LU_MO_plan['CATEGORY'] == cat, ['Area']].sum())
            if catAreaS > 0.06 or catAreaP > 0.06:
                block = pd.DataFrame([[cat, catAreaS, catAreaS-catAreaP, -catAreaP]], columns=['Zone', 'SI', 'SI+PLAN', 'PLAN'])
                Categories = Categories.append(block, ignore_index=False, verify_integrity=False, sort=False)

                
    areaMO = pd.DataFrame([[None, 'Функциональные зоны сельского поселения', areaMO_S, areaMO_P]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    SI_FZ = SI_FZ.append(areaMO, ignore_index=False, verify_integrity=False, sort=False)
    # itogo = pd.DataFrame([[None, None, SI_NP['SI'].sum(), FZ_NP['SI+PLAN'].sum()]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    itogo = pd.DataFrame([[None, None, 111111, 111111]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    SI_FZ = SI_FZ.append(itogo, ignore_index=False, verify_integrity=False, sort=False)
    mistake = pd.DataFrame([['ошибка', None, 111111, 111111]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    # mistake = pd.DataFrame([['ошибка', None, areaMO_S - SI_NP['SI'].sum(), areaMO_P - FZ_NP['SI+PLAN'].sum()]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    SI_FZ = SI_FZ.append(mistake, ignore_index=False, verify_integrity=False, sort=False)
    for group in zone_group:
        FZ_NP_A = FZ_NP.loc[FZ_NP['Zone_group'] == group]
        SI_NP_A = SI_NP.loc[SI_NP['Zone_group'] == group]
        if group != 1:
            group_name = pd.DataFrame([[None, group, SI_NP_A['SI'].sum(), FZ_NP_A['SI+PLAN'].sum()]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
            SI_FZ = SI_FZ.append(group_name, ignore_index=False, verify_integrity=False, sort=False)
        for index in zone_index:
            if index not in FZ_NP_A['Ext_Zone_Code'].tolist() and index not in SI_NP_A['Ext_Zone_Code'].tolist():
                continue
            FZ_NP_B = FZ_NP_A.loc[FZ_NP_A['Ext_Zone_Code'] == index]
            SI_NP_B = SI_NP_A.loc[SI_NP_A['Ext_Zone_Code'] == index]
            SI_FZ_A = pd.DataFrame([[index, SI_NP_B['SI'].sum(), FZ_NP_B['SI+PLAN'].sum()]], columns=['Ext_Zone_Code', 'SI', 'SI+PLAN'])
            SI_FZ_A = SI_FZ_A.fillna(0)
            SI_FZ_A = SI_FZ_A.merge(zones, left_on='Ext_Zone_Code', right_on = 'Ext_Zone_Code', how = 'left', suffixes = [None, '_1'])
            SI_FZ_A = SI_FZ_A.sort_values(by='Ext_Zone_Code')
            SI_FZ = SI_FZ.append(SI_FZ_A, ignore_index=False, verify_integrity=False, sort=False)  
    ###
    block = pd.DataFrame([[None, 'Функциональные зоны населенных пунктов', 111111, 111111]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    SI_FZ = SI_FZ.append(block, ignore_index=False, verify_integrity=False, sort=False)
    ###
    def adjust_values(si, fz, np_area, np_plan_area, np, value):
        si = si.loc[si['SI'] > 0.06]
        fz = fz.loc[fz['SI+PLAN'] > 0.06]
        si_area, fz_area = round(si['SI'].sum(), 1), round(fz['SI+PLAN'].sum(), 1)
        np_area, np_plan_area = round(np_area, 1), round(np_plan_area, 1)
        si_diff = round(np_area - si_area, 1)
        fz_diff = round(np_plan_area - fz_area, 1)

        arcpy.AddMessage(f"несоответствие {np}: {np_area}, {si_area}, {si_diff} | {np_plan_area}, {fz_area}, {fz_diff}")


        if si_diff and abs(si_diff) <= value and len(si):
            while round(si_diff, 1):

                for i in range(len(si)):
                    if round(si_diff, 1) == 0:
                        break
                    if (si['SI'].iloc[i] - 0.1) <= 0 and si['SI'].sum() != 0:
                        continue
                    if si_diff > 0:
                        si['SI'].iloc[i] += 0.1
                        si_diff -= 0.1
                    else:
                        si['SI'].iloc[i] -= 0.1
                        si_diff += 0.1

        if fz_diff and abs(fz_diff) <= value and len(fz):
            while round(fz_diff, 1):

                for i in range(len(fz)):
                    if round(fz_diff, 1) == 0:
                        break
                    if (fz['SI+PLAN'].iloc[i] - 0.1) <= 0 and fz['SI+PLAN'].sum() != 0:
                        continue
                    if fz_diff > 0:
                        fz['SI+PLAN'].iloc[i] += 0.1
                        fz_diff -= 0.1
                    else:
                        fz['SI+PLAN'].iloc[i] -= 0.1
                        fz_diff += 0.1
        return si, fz


    for name in NP['NAME']:
        FZ_NP_A = FZ_NP.loc[FZ_NP['NAME'] == name]
        SI_NP_A = SI_NP.loc[SI_NP['NAME'] == name]

        AdmeNP_SA = AdmeNP_S.loc[AdmeNP_S['NAME'] == name]
        AdmeNP_PA = AdmeNP_P.loc[AdmeNP_P['NAME'] == name]

        NP_area = AdmeNP_SA['Area'].sum()
        NP_plan_area = AdmeNP_PA['Area'].sum()

        if name == 'МО':
            NP_area = round(areaMO_S, 1) - round(float(AdmeNP.loc[AdmeNP['NOTE'] != 'планируемая', ['Area']].sum()), 1)
            NP_plan_area = round(areaMO_P, 1) - round(float(AdmeNP.loc[AdmeNP['NOTE'] != 'изменяемая', ['Area']].sum()), 1)

        if adjustValues:
            SI_NP_A, FZ_NP_A = adjust_values(SI_NP_A, FZ_NP_A, NP_area, NP_plan_area, name, adjustValues.value)
        ##
        if name == 'МО':
            MOO = pd.DataFrame([[None, 'МО', areaMO_S - float(AdmeNP.loc[AdmeNP['NOTE'] != 'планируемая', ['Area']].sum()), areaMO_P - float(AdmeNP.loc[AdmeNP['NOTE'] != 'изменяемая', ['Area']].sum())]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
            SI_FZ = SI_FZ.append(MOO, ignore_index=False, verify_integrity=False, sort=False)
        ##

        np_name = pd.DataFrame([[None, name, AdmeNP_SA['Area'].sum(), AdmeNP_PA['Area'].sum()]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
        SI_FZ = SI_FZ.append(np_name, ignore_index=False, verify_integrity=False, sort=False)
        ###
        for group in zone_group:
            FZ_NP_B = FZ_NP_A.loc[FZ_NP_A['Zone_group'] == group]
            SI_NP_B = SI_NP_A.loc[SI_NP_A['Zone_group'] == group]
            if group != 1:
                group_name = pd.DataFrame([[None, group, SI_NP_B['SI'].sum(), FZ_NP_B['SI+PLAN'].sum()]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
                SI_FZ = SI_FZ.append(group_name, ignore_index=False, verify_integrity=False, sort=False)
            SI_FZ_A = SI_NP_B.merge(FZ_NP_B, left_on='Ext_Zone_Code', right_on = 'Ext_Zone_Code', how = 'outer', suffixes = [None, '_1'])
            SI_FZ_A = SI_FZ_A.fillna(0)
            SI_FZ_A = SI_FZ_A.merge(zones, left_on='Ext_Zone_Code', right_on = 'Ext_Zone_Code', how = 'left')
            SI_FZ_A = SI_FZ_A.sort_values(by='Ext_Zone_Code')
            SI_FZ = SI_FZ.append(SI_FZ_A, ignore_index=False, verify_integrity=False, sort=False)
        itogo = pd.DataFrame([['ИТОГО', None, SI_NP_A['SI'].sum(), FZ_NP_A['SI+PLAN'].sum()]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])

        SI_FZ = SI_FZ.append(itogo, ignore_index=False, verify_integrity=False, sort=False)

        mistake = pd.DataFrame([['ошибка', None, 111111, 111111]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
        SI_FZ = SI_FZ.append(mistake, ignore_index=False, verify_integrity=False, sort=False)


    SI_FZ = SI_FZ.loc[(SI_FZ['SI'] > 0.06) | (SI_FZ['SI+PLAN'] > 0.06) != 0]

    #SI_FZ = SI_FZ.round(1)

    SI_FZ['PLAN'] = SI_FZ['SI+PLAN'] - SI_FZ['SI']
    Balance['PLAN'] = Balance['SI+PLAN'] - Balance['SI']
    #SI_FZ = SI_FZ.loc[(SI_FZ['SI'] != 0) | (SI_FZ['SI+PLAN'] != 0) != 0]
    Balance = Balance.append(Categories, ignore_index=False, verify_integrity=False, sort=False)
    block = pd.DataFrame([['2.', 'Функциональные зоны', None, None]], columns=['Ext_Zone_Code', 'Zone', 'SI', 'SI+PLAN'])
    Balance = Balance.append(block, ignore_index=False, verify_integrity=False, sort=False)
    Balance = Balance.append(SI_FZ, ignore_index=False, verify_integrity=False, sort=False)
    Balance = Balance.round(1)


    Balance['B'] = None
    Balance['D'] = None
    Balance = Balance.rename({'Ext_Zone_Code': 'A', 'Zone': 'C', 'SI': 'E', 'SI+PLAN': 'F', 'PLAN': 'G'}, axis = 1)
    Balance = Balance[['A', 'B', 'C', 'D', 'E', 'F', 'G']+lu_columns+['ИТОГО']]
    Balance = Balance.where(pd.notnull(Balance), None)

    zones_dict = {i:j for i, j in zip(zones['Ext_Zone_Code'].to_list(), zones['Zone_group'].to_list())}
    cont = 0
    fz = 0
    fz_np_ind = Balance['C'].to_list().index('Функциональные зоны населенных пунктов') + 2
    b_end = len(Balance['C'].to_list()) + 1
    fz_E = '='
    fz_F = '='
    itogo_E = '='
    itogo_F = '='
    np_ind_Dict = {i: [0, 0] for i in NP['NAME'].to_list()}
    # arcpy.AddMessage(str(np_ind_Dict))
    # Balance.to_excel(outputXLS.valueAsText)
    for i in range(len(Balance)):
        n = i+2

        if Balance['C'].iloc[i] == 'Функциональные зоны населенных пунктов':
            Balance['E'].iloc[i] = None
            Balance['F'].iloc[i] = None
            Balance['G'].iloc[i] = None
            fz = 2
            Balance['E'].iloc[fz_ind] = fz_E[:-1]
            Balance['F'].iloc[fz_ind] = fz_F[:-1]
            Balance['E'].iloc[fz_ind+1] = f'=E{fz_ind+2}-E{fz_ind-1+2}'
            Balance['F'].iloc[fz_ind+1] = f'=F{fz_ind+2}-F{fz_ind-1+2}'

            continue
        elif fz == 2:
            if Balance['C'].iloc[i] == 'МО':
                Balance['E'].iloc[i] = f'=E2-E{NP_ind}'
                Balance['F'].iloc[i] = f'=F2-F{NP_ind}'
                current_np_ind = n
                current_np = Balance['C'].iloc[i]
                np_ind_Dict[current_np][0] = n

            elif Balance['A'].iloc[i] == 'ИТОГО':
                Balance['E'].iloc[i] = itogo_E[:-1]
                Balance['F'].iloc[i] = itogo_F[:-1]
                itogo_E = '='
                itogo_F = '='
                Balance['E'].iloc[i+1] = f'=E{n}-E{current_np_ind}'
                Balance['F'].iloc[i+1] = f'=F{n}-F{current_np_ind}'

            elif Balance['A'].iloc[i] == 'ошибка':
                np_ind_Dict[current_np][1] = n
                pass

            elif Balance['A'].iloc[i] == None:
                if Balance['C'].iloc[i] in zone_group:
                    itogo_E += f'E{n}+'
                    itogo_F += f'F{n}+'
                    nn = 0
                    for j in Balance['A'][i+1:].to_list():
                        if j != None: 
                            if j not in ['ИТОГО', 'ошибка']:
                                if zones_dict[j] != 1:
                                    nn += 1
                        else:
                            break

                    #nn = len(Balance['A'].to_list()[i+1:len(Balance['A'][:i+1])+Balance['A'][i+1:].to_list().index(None)])
                    Balance['E'].iloc[i] = f'=SUM(E{n+1}:E{n+nn})'
                    Balance['F'].iloc[i] = f'=SUM(F{n+1}:F{n+nn})'
                else:
                    # Ссылаемся на площади нп из таблицы категорий
                    current_np_ind = n
                    current_np = Balance['C'].iloc[i]
                    np_ind_Dict[current_np][0] = n
                    np_pos = Balance['C'].to_list().index(Balance['C'].iloc[i]) + 2
                    Balance['E'].iloc[i] = f'=SUM(E{n+1}:E{n+nn})'
                    Balance['F'].iloc[i] = f'=SUM(F{n+1}:F{n+nn})'
                    Balance['E'].iloc[i] = f'=E{np_pos}'
                    Balance['F'].iloc[i] = f'=F{np_pos}'
            elif zones_dict[Balance['A'].iloc[i]] == 1:
                itogo_E += f'E{n}+'
                itogo_F += f'F{n}+'
            else:
                pass

        elif fz == 1:
            if Balance['C'].iloc[i] != None:
                if Balance['C'].iloc[i] != 'Функциональные зоны сельского поселения':
                    if Balance['A'].iloc[i] != None:
                        if zones_dict[Balance['A'].iloc[i]] == 1:
                            fz_E += f'E{n}+'
                            fz_F += f'F{n}+'
                        else:
                            pass
                        Balance['E'].iloc[i] = f'=SUMIFS(E{fz_np_ind}:E{b_end},A{fz_np_ind}:A{b_end},A{n})'
                        Balance['F'].iloc[i] = f'=SUMIFS(F{fz_np_ind}:F{b_end},A{fz_np_ind}:A{b_end},A{n})'
                        Balance['G'].iloc[i] = f'=F{n}-E{n}'
                        continue
                    else:
                        fz_E += f'E{n}+'
                        fz_F += f'F{n}+'
                        nn = 0
                        for j in Balance['A'][i+1:].to_list():
                            if j != None: 
                                if zones_dict[j] != 1:
                                    nn += 1
                            else:
                                break

                        #nn = len(Balance['A'].to_list()[i+1:len(Balance['A'][:i+1])+Balance['A'][i+1:].to_list().index(None)])
                        Balance['E'].iloc[i] = f'=SUM(E{n+1}:E{n+nn})'
                        Balance['F'].iloc[i] = f'=SUM(F{n+1}:F{n+nn})'
                else:
                    Balance['C'].iloc[i] = mo_type
        elif Balance['C'].iloc[i] == 'Функциональные зоны':
            fz = 1
            fz_ind = i + 2
            continue

        elif Balance['E'].iloc[i] == None:
            continue

        elif Balance['A'].iloc[i] == '1.1':
            nn = len(Balance['A'].to_list()[i:Balance['A'].to_list().index('1.2')])-1
            NP_end = nn
            NP_ind = n
            Balance['E'].iloc[i] = f'=SUM(E{n+1}:E{n+nn})'
            Balance['F'].iloc[i] = f'=SUM(F{n+1}:F{n+nn})'
        elif Balance['A'].iloc[i] == '1.2':
            cont = 1
            CAT_ind = n
        elif Balance['A'].iloc[i] == 'ИТОГО' and Balance['A'].iloc[i-1][0] == '1':

            Balance['E'].iloc[i] = f'=SUM(E{NP_ind},E{CAT_ind}:E{n-1})'
            Balance['G'].iloc[i] = f'=SUM(G{NP_ind},G{CAT_ind}:G{n-1})'
            Balance['F'].iloc[i] = f'=E{n}+G{n}'
            continue
        elif Balance['A'].iloc[i] == 'ошибка':
            cont = 0
            # Cчитаем ошибку по категориям
            cell_E = f'E{n - 1}-E2'
            cell_G = f'G{n - 1}-G2'
            Balance['F'].iloc[i] = f'=E{n}+G{n}'
            if Balance['C'].iloc[i + 1] == 'Иные категории земель в границах населенных пунктов':
                cell_E += f'-E{n + 1}'
                cell_G += f'-G{n + 1}'

            Balance['E'].iloc[i] = f'=ROUND({cell_E}, 1)'
            Balance['G'].iloc[i] = f'=ROUND({cell_G}, 1)'
            continue
        elif Balance['A'].iloc[i] == '*':
            nn = len(Balance['A'].to_list()[i:Balance['A'].to_list().index('2.')])-1
            Balance['E'].iloc[i] = f'=SUM(E{n+1}:E{n+nn})'
            Balance['F'].iloc[i] = f'=SUM(F{n+1}:F{n+nn})'
        else:
            pass

        if cont == 0:
            Balance['G'].iloc[i] = f'=F{n}-E{n}'
        elif cont == 1:
            Balance['F'].iloc[i] = f'=E{n}+G{n}'
        else:
            continue

    Balance = Balance.rename({'A': '', 'B': '№ п/п', 'C': 'Наименование показателя', 'D': 'Единица измерения', 'E': 'Современное состояние, га', 'F': 'Современное состояние + планируемые, га', 'G': 'Планируемые, га'}, axis = 1)
    Balance = Balance[['', '№ п/п', 'Наименование показателя', 'Единица измерения', 'Современное состояние, га', 'Современное состояние + планируемые, га', 'Планируемые, га']+lu_columns+['ИТОГО']]
    writer = pd.ExcelWriter(outputXLS.valueAsText)
    Balance.to_excel(writer, engine='openpyxl', sheet_name='Balance', index=False) # openpyxl
    # Работа со стилями
    def set_border(ws, side=None, blank=True):
        wb = ws._parent
        side = side if side else Side(border_style='thin', color='000000')
        for cell in ws._cells.values():
            cell.border = Border(top=side, bottom=side, left=side, right=side)
            cell.alignment =  cell.alignment.copy(wrapText=True)
            cell.number_format = '0.0'
        if blank:
            white = Side(border_style='thin', color='FFFFFF')
            wb._borders.append(Border(top=white, bottom=white, left=white, right=white))
            wb._cell_styles[0].borderId = len(wb._borders) - 1
    # Открываем книгу через openpyxl
    workbook = writer.book
    worksheet = writer.sheets['Balance']

    def col_letter(n):
        return chr(ord('@')+n)

    additional_columns = [col_letter(7 + i) for i in range(1, len(lu_columns) + 1)]
    itogo_column = col_letter(7 + len(lu_columns) + 1)

    # Задаем ширину столбцов
    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 9
    worksheet.column_dimensions['C'].width = 70
    worksheet.column_dimensions['D'].width = 11
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15            

    for letter in additional_columns:
        worksheet.column_dimensions[letter].width = 20

    worksheet.column_dimensions[itogo_column].width = 9

    # Создаем границы ячеек

    side = Side(border_style='thin', color='000000')
    set_border(worksheet, side, False)

    # Создаем толстые внешние границы
    def thick_border(ws, cell_range, side=None):
        wb = ws._parent
        ws = ws[cell_range]
        side = side if side else Side(border_style='thin', color='000000')
        thick_side = Side(border_style='thick', color='000000')
        top = side
        bottom = side
        left = side
        right = side
        for i, row in enumerate(ws):
            if i == 0:
                top = thick_side
                bottom = side
            elif i == len(ws)-1:
                bottom = thick_side
                top = side
            else:
                top = side
                bottom = side
            for j, cell in enumerate(row):
                if j == 0:
                    left = thick_side
                    right = side
                elif j == len(row) - 1:
                    right = thick_side
                    left = side
                else:
                    left = side
                    right = side
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    thick_border(worksheet, (f'A1:G{fz_ind-1}'))
    thick_border(worksheet, (f'A{fz_ind}:G{fz_np_ind-1}'))

    # Создаем условное форматирование для отображения ошибок
    redFill = PatternFill(bgColor='FEB4B4') # fill_type='solid', 

    # (max_row, max_col) = Balance.shape
    worksheet.conditional_formatting.add(f'$A${fz_ind+3}:$G${fz_ind+3}', FormulaRule(formula=[f'IF(COUNTIFS($A${fz_np_ind}:$A${worksheet.max_row}, "ошибка", $E${fz_np_ind}:$E${worksheet.max_row}, "<>0")+COUNTIFS($A${fz_np_ind}:$A${worksheet.max_row}, "ошибка", $F${fz_np_ind}:$F${worksheet.max_row}, "<>0"), 1)'], fill=redFill))
    for i, ind in np_ind_Dict.items():
        thick_border(worksheet, (f'A{ind[0]}:G{ind[1]}'))
        worksheet.conditional_formatting.add(f'$A${ind[0]}:$G${ind[1]}', FormulaRule(formula=[f'IF(AND($A${ind[1]}="ошибка", OR($E${ind[1]}<>0, $F${ind[1]}<>0)), 1)'], fill=redFill))

    # Заливаем категории
    categoryFill = PatternFill(fill_type = 'solid', start_color='FCD5B4', end_color='FCD5B4')
    for row in worksheet['A1':f'G{fz_ind-1}']:
        for cell in row:
            cell.fill = categoryFill

    # Заливаем ФЗ
    fzFill = PatternFill(fill_type = 'solid', start_color='B8CCE4', end_color='B8CCE4')
    for row in worksheet[f'A{fz_ind}':f'G{fz_np_ind-1}']:
        for cell in row:
            cell.fill = fzFill


    # Добавляем жирный текст Категории
    for row in worksheet[f'A2':f'G{fz_ind-1}']:
        if row[0].value in ['ИТОГО', 'ошибка', '1.', '*']:

            if row[0].value == 'ошибка':
                cell_font = Font(bold=True, color='C00000')
            else:
                cell_font = Font(bold=True)
            for cell in row:
                cell.font = cell_font

    # Добавляем жирный текст ФЗ
    for row in worksheet[f'A{fz_ind}':f'G{worksheet.max_row}']:
        if not row[0].value or row[0].value in ['ИТОГО', 'ошибка'] or zones_dict.get(row[0].value) == 1:

            if row[0].value == 'ошибка':
                cell_font = Font(bold=True, color='C00000')
            else:
                cell_font = Font(bold=True)
            for cell in row:
                cell.font = cell_font

    # Заполняем номера НП и зон

    for row in worksheet[f'A1:G{fz_ind-1}']:
        if row[0].value in ['*'] or '.' in row[0].value: 
            row[1].value = row[0].value
            row[0].value = None


    next_value = None
    group_n = 1
    for row in worksheet[f'A{fz_ind}':f'G{fz_np_ind - 1}']:
        if next_value:
            row[1].value = next_value
            next_value = None
            continue
        if row[2].value == 'Функциональные зоны':
            row[0].value = None
            row[1].value = '2.'
            next_value = '2.1'
        if row[2].value in zone_group or zones_dict.get(row[0].value) == 1:
            row[1].value = '2.1.' + str(group_n)
            group_n += 1


    np_n = 1
    group_n = 1
    for row in worksheet[f'A{fz_np_ind}':f'G{worksheet.max_row}']:
        if row[2].value == 'Функциональные зоны населенных пунктов':
            row[1].value = '2.2'
        if row[2].value in np_ind_Dict:
            row[1].value = '2.2.' + str(np_n)
            current_np_value = row[1].value
            np_n += 1
            group_n = 1
        if row[2].value in zone_group or zones_dict.get(row[0].value) == 1:
            row[1].value = current_np_value + '.' + str(group_n)
            group_n += 1

    for row in worksheet:
        if row[3].value is not None and row[3].value != 'Единица измерения':
            row[3].value = 'га'
            

    for row in worksheet[f'A1:{col_letter(worksheet.max_column)}{fz_ind-1}']:
        if type(row[1].value) != str:
            continue
        if row[1].value[:2].startswith('1.') and len(row[1].value) > 2:
            row[-1].value = f'=SUM(H{row[0].row}:{col_letter(worksheet.max_column-1)}{row[0].row})'
        if row[0].value == 'ИТОГО':
            itogo_ind = row[0].row
            for i in range(7, worksheet.max_column):
                row[i].value = f'=SUM({col_letter(i+1)}3:{col_letter(i+1)}{row[0].row-1})'

    for row in worksheet[f'A1:{col_letter(worksheet.max_column)}{fz_ind-1}']:
        if type(row[1].value) != str:
            continue
        if row[1].value[:2].startswith('1.') and len(row[1].value) > 2 and row[1].value != '1.1':
            if row[2].value in lu_columns:
                lu_index = lu_columns.index(row[2].value) + 2
                row[6].value = f'={col_letter(6+lu_index)}{itogo_ind}-{col_letter(worksheet.max_column)}{row[0].row}'
            else:
                row[6].value = f'=-{col_letter(worksheet.max_column)}{row[0].row}'
        elif row[1].value == '1.1':
            lu_np_ind = row[0].row
            lu_np = row[2].value
            
        if row[0].value == 'ошибка':
            if lu_np in lu_columns:
                lu_index = lu_columns.index(lu_np) + 8
                row[lu_index-1].value = f'={col_letter(lu_index)}{itogo_ind}-{col_letter(worksheet.max_column)}{lu_np_ind}-G{lu_np_ind}'
        
    # Конец работы со стилями
    #LU_PIV.to_excel(writer, engine='openpyxl', sheet_name='LandUse', index=True)
    LU_int_minus.to_excel(writer, engine='openpyxl', sheet_name='ПересеченияПланКатегорий', index=True)
    writer.close()

    if deleteMemory.value:
        arcpy.management.Delete(output_db)

if __name__ == '__main__':
    execute()
